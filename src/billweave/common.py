#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
common.py —— 财务数据共享读取层(calc_scenario.py / ledger.py 共用)

职责:
  1. 递归扫描 balance/、bills/、debt/ 下所有支持格式的文件(.csv/.xlsx/.xlsm/.pdf)
  2. 读取并标准化为统一交易/余额/债务记录
  3. 处理真实账单的常见"野格式":
     - CSV 编码探测:utf-8-sig → gb18030 → latin-1(支付宝导出为 GBK)
     - 元信息行跳过:微信/支付宝账单前部有标题、昵称、说明等行,用关键词定位真正的表头
     - PDF:优先 extract_tables();无表格线时按坐标分列(招行等银行流水常见排版);
       两者都失败再退化为按空格切分文本行
     - 日期/金额标准化:datetime、2026/9/1、2026年9月1日、20260830、千分位、货币符号
     - 收支判断:优先"收/支"列,其次金额正负号

字段说明(标准化后的交易 dict):
  日期/平台/类型/项目/金额/币种/状态/对方(交易对手,原字段名"备注"保留兼容)/
  支付方式(微信/支付宝记录的付款渠道,如"招商银行储蓄卡(1234)"/"零钱"/"来源")
"""

import csv
import os
import re
from datetime import date, datetime, timedelta

SUPPORTED_EXTS = (".csv", ".xlsx", ".xlsm", ".pdf")

# 表头定位关键词:真实账单表头行通常包含其中多个词
HEADER_KEYWORDS = [
    "交易时间", "记账日期", "交易日期", "日期", "时间",
    "金额", "交易金额", "发生额",
    "交易对方", "商品", "摘要", "用途", "项目",
    "收/支", "收支", "类型", "交易类型",
    "币种", "货币", "账户", "平台", "状态",
]

# 列名别名:内部字段 -> 可能的列名(优先级从高到低)
ALIASES = {
    "日期":   ["交易时间", "交易日期", "记账日期", "交易创建时间", "数据日期", "日期", "时间"],
    "金额":   ["金额(元)", "金额（元）", "交易金额", "金额", "发生额", "收支金额", "amount"],
    "类型":   ["收/支", "收/支类型", "收支类型", "交易类型", "类型", "借贷", "收支"],
    "项目":   ["商品", "商品名称", "交易摘要", "摘要", "项目", "用途", "名称", "name"],
    "对方":   ["交易对方", "对手信息", "对方", "商户", "债权人", "creditor"],
    "支付方式": ["支付方式", "付款方式", "收/付款方式"],
    "平台":   ["交易渠道", "渠道", "平台", "来源"],
    "状态":   ["当前状态", "交易状态", "状态", "资金状态", "还款状态"],
    "账户":   ["账户", "账号", "卡号", "户名", "account"],
    "币种":   ["币种", "货币", "currency"],
    "受限原因": ["受限原因", "冻结原因", "restricted", "备注", "说明"],
    "还款日期": ["还款日期", "还款日", "到期日"],
}

# 支出/收入类型关键词(兼容各平台"收/支"列值)
EXPENSE_WORDS = ("支出", "expense", "支取", "借方", "付款", "消费", "转出", "还款")
INCOME_WORDS = ("收入", "income", "存入", "贷方", "退款", "收款", "转入")
# 中性交易:支付宝"不计收支"(余额宝转入转出、信用卡/花呗还款等),既不收入也不支出,不录入账本
NEUTRAL_WORDS = ("不计收支", "不计", "中性", "非收支")
# 付款渠道显示为银行卡的关键词(用于识别微信/支付宝的"跨平台结算"候选)
BANK_CARD_WORDS = ("银行卡", "储蓄卡", "信用卡", "借记卡")


def detect_encoding(path):
    for enc in ("utf-8-sig", "gb18030", "latin-1"):
        try:
            with open(path, encoding=enc) as f:
                f.read(4096)
            return enc
        except (UnicodeDecodeError, OSError):
            continue
    return "latin-1"


def to_number(v):
    """金额/数字转 float;失败返回 None。支持千分位、货币符号。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("¥", "").replace("￥", "").replace("$", "").strip()
    if s in ("", "-", "--"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def normalize_date(v):
    """多种日期格式 → YYYY-MM-DD;失败返回空串。"""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return ""
    m = re.match(r"^(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})日?", s)
    if m:
        try:
            return "%04d-%02d-%02d" % (int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return ""
    if re.match(r"^\d{8}$", s):
        return "%s-%s-%s" % (s[0:4], s[4:6], s[6:8])
    return ""


def clean(v):
    return str(v).replace("\n", "").replace("\r", "").strip() if v is not None else ""


def is_bank_card_payment(text):
    """判断一段付款方式/摘要文字是否指向银行卡付款(用于'跨平台结算'去重候选识别)。"""
    t = clean(text)
    return any(w in t for w in BANK_CARD_WORDS)


def _find_header_idx(rows, header_score_fn, min_score=2):
    """在 rows 前 40 行内找得分最高的表头行;得分不足返回 None。"""
    best = (0, None)
    for i, r in enumerate(rows[:40]):
        score = header_score_fn(r)
        if score > best[0]:
            best = (score, i)
    if best[0] >= min_score:
        return best[1]
    return None


def _header_score(row):
    joined = "".join(clean(c) for c in row)
    return sum(1 for k in HEADER_KEYWORDS if k in joined)


def read_csv(path):
    """读 CSV:编码探测 + 元信息行跳过。返回 (表头列表, 行列表[dict])。"""
    enc = detect_encoding(path)
    with open(path, encoding=enc, newline="") as f:
        raw = list(csv.reader(f))
    idx = _find_header_idx(raw, _header_score)
    if idx is None:
        return [], []
    headers = [clean(c) for c in raw[idx]]
    out = []
    for r in raw[idx + 1:]:
        if all(clean(c) == "" for c in r):
            continue
        row = {}
        for j, h in enumerate(headers):
            if h and j < len(r):
                row[h] = r[j]
        out.append(row)
    return headers, out


def read_excel(path):
    """读 Excel:表头关键词定位(微信账单表头在第18行)。返回 (表头, 行[dict])。"""
    try:
        import openpyxl
    except ImportError:
        return [], []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = next((sh for sh in wb.worksheets if sh.max_row and sh.max_column), None)
    if ws is None:
        return [], []
    rows = list(ws.iter_rows(values_only=True))
    idx = _find_header_idx(rows, _header_score)
    if idx is None:
        return [], []
    headers = [clean(c) for c in rows[idx]]
    out = []
    for r in rows[idx + 1:]:
        if all(c is None or clean(c) == "" for c in r):
            continue
        row = {}
        for j, h in enumerate(headers):
            if h and j < len(r):
                row[h] = r[j]
        out.append(row)
    return headers, out


def _group_words_into_lines(words, tol=2.5):
    """把 pdfplumber 逐词坐标按行归拢(top 坐标接近的算一行),行内按 x0 排序。"""
    lines = []
    for w in sorted(words, key=lambda w: (w["top"], w["x0"])):
        placed = False
        for line in lines:
            if abs(line[0]["top"] - w["top"]) <= tol:
                line.append(w)
                placed = True
                break
        if not placed:
            lines.append([w])
    for line in lines:
        line.sort(key=lambda w: w["x0"])
    lines.sort(key=lambda line: line[0]["top"])
    return lines


def _read_pdf_page_by_columns(page):
    """按表头 x 坐标切列,应对无网格线的银行流水 PDF(如招行:记账日期/货币/交易金额/
    联机余额/交易摘要/对手信息,列与列之间没有表格线,只能靠对齐)。
    比按空格数切分更稳:能正确处理摘要/对手信息内部本身带空格的情况。
    失败(找不到表头/没有数据行)返回 (None, [])。"""
    try:
        words = page.extract_words()
    except Exception:
        return None, []
    if not words:
        return None, []
    lines = _group_words_into_lines(words)

    best_idx, best_score = None, 0
    for i, line in enumerate(lines[:40]):
        joined = "".join(clean(w["text"]) for w in line)
        score = _header_score([joined])
        if score > best_score:
            best_score, best_idx = score, i
    if best_idx is None or best_score < 2:
        return None, []

    header_words = lines[best_idx]
    headers = [clean(w["text"]) for w in header_words]
    bounds = [w["x0"] for w in header_words]

    def bucket(x0):
        col = 0
        for j, b in enumerate(bounds):
            if x0 >= b - 1:
                col = j
        return col

    out = []
    for line in lines[best_idx + 1:]:
        first_text = clean(line[0]["text"])
        if not normalize_date(first_text):
            continue  # 非数据行(表头续行/英文小标题/页脚说明等),跳过
        cells = [""] * len(headers)
        for w in line:
            col = bucket(w["x0"])
            cells[col] = (cells[col] + " " + clean(w["text"])).strip()
        row = {h: cells[j] for j, h in enumerate(headers) if h}
        out.append(row)
    if not out:
        return None, []
    return headers, out


def read_pdf(path):
    """读 PDF,依次尝试三种策略:
    1. extract_tables():有网格线的表格
    2. 按坐标分列:无网格线但版式对齐的银行流水(如招行),能正确拆出独立的"对手信息"列
    3. 按空格切分文本行:前两种都失败时的兜底,摘要与对手信息会被合并
    返回 (表头, 行[dict])。"""
    try:
        import pdfplumber
    except ImportError:
        return [], []

    rows = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for t in page.extract_tables():
                for r in t:
                    rows.append([clean(c) for c in r])
    if rows:
        idx = _find_header_idx(rows, _header_score)
        headers = [clean(c) for c in rows[idx]] if idx is not None else []
        out = []
        for r in (rows[idx + 1:] if idx is not None else rows):
            if all(clean(c) == "" for c in r):
                continue
            if headers:
                row = {}
                for j, h in enumerate(headers):
                    if h and j < len(r):
                        row[h] = r[j]
                out.append(row)
        if out:
            return headers, out

    headers_by_col, rows_by_col = None, []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            h, page_rows = _read_pdf_page_by_columns(page)
            if h:
                headers_by_col = headers_by_col or h
                rows_by_col.extend(page_rows)
    if rows_by_col:
        return headers_by_col or [], rows_by_col

    # 兜底:按空格切分文本行(摘要与对手信息会被合并成一个字段,无法单独识别对手方)
    rows = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for line in (page.extract_text() or "").splitlines():
                line = line.strip()
                if not line:
                    continue
                if re.search(r"[：:]", line) and not re.match(r"^\d{4}[-/]", line):
                    continue
                parts = line.split()
                if len(parts) >= 5 and re.match(r"^\d{4}-\d{2}-\d{2}$", parts[0]):
                    rows.append([parts[0], parts[1], parts[2], parts[3],
                                 " ".join(parts[4:])])
    if not rows:
        return [], []
    idx = _find_header_idx(rows, _header_score)
    headers = [clean(c) for c in rows[idx]] if idx is not None else []
    out = []
    for r in (rows[idx + 1:] if idx is not None else rows):
        if all(clean(c) == "" for c in r):
            continue
        if headers:
            row = {}
            for j, h in enumerate(headers):
                if h and j < len(r):
                    row[h] = r[j]
            out.append(row)
        else:
            out.append({"日期": r[0], "币种": r[1], "交易金额": r[2], "交易摘要": " ".join(r[4:])})
    return headers, out


def read_table(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return read_csv(path)
    if ext in (".xlsx", ".xlsm"):
        return read_excel(path)
    if ext == ".pdf":
        return read_pdf(path)
    return [], []


def find_files(root, exts=SUPPORTED_EXTS):
    """递归扫描目录下所有支持格式的文件。"""
    files = []
    if not os.path.isdir(root):
        return files
    for dirpath, _, fnames in os.walk(root):
        for fn in sorted(fnames):
            if fn.lower().endswith(exts):
                files.append(os.path.join(dirpath, fn))
    return sorted(files)


def get_field(row, field):
    for alias in ALIASES.get(field, [field]):
        v = row.get(alias)
        if v is not None and clean(v) != "":
            return v
    return ""


def _file_amount_signed(rows):
    """判断该账单文件金额列是否带符号(存在负数)——银行流水(如招商银行)金额带符号,
    微信/支付宝金额全为正。用于类型列读不到时的兜底方式:
    带符号→按金额符号判,全为正→归中性。"""
    for r in rows:
        a = to_number(get_field(r, "金额"))
        if a is not None and a < 0:
            return True
    return False


def classify_income_expense(row, amount_signed=False):
    """判断收支:优先'收/支'列值。类型列读不到时,若金额列带符号(有正有负,如银行流水)则按金额符号判;
    金额列全为正(如微信/支付宝)则归中性(不计收支)。返回 income/expense/neutral。
    注: '交易关闭'(未成交)交由 ledger 去重阶段按优先级处理,不在此提前标中性,
    以免掩盖原始收支方向(避免退款2匹配不到)。"""
    typ = clean(get_field(row, "类型")).lower()
    if any(w in typ for w in NEUTRAL_WORDS):
        return "neutral"  # 不计收支:先于收支判断,避免被误判
    if any(w in typ for w in EXPENSE_WORDS):
        return "expense"
    if any(w in typ for w in INCOME_WORDS):
        return "income"
    # 类型列无法判断 → 按金额列符号分布决策
    if amount_signed:
        amt = to_number(get_field(row, "金额"))
        if amt is not None and amt != 0:
            return "expense" if amt < 0 else "income"
    return "neutral"


def extract_transaction(path, row, platform, amount_signed=False):
    """把一行原始记录标准化为统一交易 dict。解析失败返回 None。"""
    date_s = normalize_date(get_field(row, "日期"))
    amt = to_number(get_field(row, "金额"))
    if not date_s or amt is None:
        return None
    # 项目清洗:商品列为空/占位符/像单号(商户单号xxx 或超长无意义串)时,回退用交易对方
    item = clean(get_field(row, "项目"))
    peer = clean(get_field(row, "对方"))
    if (not item or item in ("/", "-", "--") or item.startswith("商户单号")
            or re.fullmatch(r"[\dA-Za-z/.-]{12,}", item)):
        item = peer
    pay_method = clean(get_field(row, "支付方式"))
    ie = classify_income_expense(row, amount_signed)
    base = {
        "日期": date_s, "时间": clean(get_field(row, "日期")), "平台": platform,
        "项目": item, "金额": amt,
        "币种": clean(get_field(row, "币种")).upper() or "CNY",
        "状态": clean(get_field(row, "状态")),
        "对方": peer, "备注": peer, "支付方式": pay_method,
        "来源": os.path.basename(path),
    }
    if ie == "neutral":
        # 不计收支(余额宝/信用卡还款等):不录入账本,由 ledger.py 剔除
        base["类型"] = "neutral"
        return base
    if ie == "expense" and amt > 0:
        amt = -abs(amt)  # 支出列金额为正,统一为负(招行金额本身带符号,不动)
    elif ie == "income" and amt < 0:
        amt = abs(amt)
    base["类型"] = ie
    base["金额"] = amt
    return base


def load_transactions(finance_dir, subdir="bills", platform_hint=None):
    """递归读取某目录下所有账单,返回统一交易列表。
    platform 取文件所在目录名(如微信/支付宝/招商银行)。"""
    root = os.path.join(finance_dir, subdir)
    txs = []
    for path in find_files(root):
        headers, rows = read_table(path)
        if not rows:
            continue
        rel = os.path.relpath(path, root)
        # 平台:文件所在的第一级子目录名(如 bills/微信/2026.08.30/xxx → 微信)
        parts = rel.split(os.sep)
        platform = parts[0] if len(parts) > 1 and parts[0] not in ("", ".") else platform_hint or "未知平台"
        amount_signed = _file_amount_signed(rows)
        for r in rows:
            tx = extract_transaction(path, r, platform, amount_signed)
            if tx:
                txs.append(tx)
    return txs


def try_read_wide(path):
    """横表检测:表头形如 [日期, 账户1, 账户2, ...],数据行 [日期, 余额1, ...]。
    返回 [(账户, 金额, 日期), ...];失败返回 []。"""
    if path.lower().endswith((".xlsx", ".xlsm")):
        try:
            import openpyxl
            ws = openpyxl.load_workbook(path, data_only=True, read_only=True).worksheets[0]
            rows = list(ws.iter_rows(values_only=True))
        except Exception:
            return []
    else:
        try:
            with open(path, encoding=detect_encoding(path), newline="") as f:
                rows = list(csv.reader(f))
        except Exception:
            return []
    out = []
    for i, r in enumerate(rows[:10]):
        cells = [clean(c) for c in r]
        if len(cells) < 3 or cells[0] not in ("日期", "date", "时间"):
            continue
        has_date = False
        for j in range(i + 1, min(i + 2, len(rows))):
            if normalize_date(rows[j][0]):
                has_date = True
                break
        if not has_date:
            continue
        for k in range(1, len(cells)):
            if not cells[k]:
                continue
            for j in range(i + 1, len(rows)):
                d = normalize_date(rows[j][0])
                if not d:
                    continue
                amt = to_number(rows[j][k]) if k < len(rows[j]) else None
                if amt is not None:
                    out.append((cells[k], amt, d))
        return out
    return []


def load_balances(finance_dir):
    """读 balance/ 下所有余额快照(竖表/横表都支持)。返回余额行列表。"""
    rows = []
    for path in find_files(os.path.join(finance_dir, "balance")):
        headers, raw = read_table(path)
        if raw:
            for r in raw:
                amt = to_number(get_field(r, "金额"))
                if amt is None:
                    continue
                rows.append({
                    "账户": get_field(r, "账户") or os.path.splitext(os.path.basename(path))[0],
                    "金额": amt,
                    "币种": (get_field(r, "币种") or "CNY").upper(),
                    "数据日期": normalize_date(get_field(r, "日期")) or "",
                    "受限原因": get_field(r, "受限原因"),
                    "来源": os.path.basename(path),
                })
        else:
            # 横表:日期 + 账户名列
            for acct, amt, d in try_read_wide(path):
                rows.append({
                    "账户": acct, "金额": amt, "币种": "CNY",
                    "数据日期": d, "受限原因": "", "来源": os.path.basename(path),
                })
    # 统一口径: 按账户取最新日期快照（历史行忽略不计）——所有消费方共用同一规则，
    # 与 overview 的余额汇总、weekly 的"长期未更新"判断一致；日期更大才覆盖，
    # 空日期行保留（有日期行会覆盖它）。
    latest = {}
    for b in rows:
        acct = b["账户"]
        prev = latest.get(acct)
        if prev is None or (b["数据日期"] and b["数据日期"] > prev["数据日期"]):
            latest[acct] = b
    return list(latest.values())


def load_debts(finance_dir):
    """读 debt/ 下所有债务。返回债务行列表。"""
    rows = []
    for path in find_files(os.path.join(finance_dir, "debt")):
        headers, raw = read_table(path)
        for r in raw:
            amt = to_number(get_field(r, "金额"))
            if amt is None:
                continue
            rows.append({
                "债权人": get_field(r, "对方") or get_field(r, "账户") or os.path.splitext(os.path.basename(path))[0],
                "金额": amt,
                "币种": (get_field(r, "币种") or "CNY").upper(),
                "数据日期": normalize_date(get_field(r, "日期")) or "",
                "状态": get_field(r, "状态"),
                "来源": os.path.basename(path),
            })
    return rows
